from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.drawing.image import Image
import os
#------------------------------------------------
#REPORTE MENSUAL
#------------------------------------------------
def exportar_reporte_mensual_excel(
        empresa_nombre,
        nombre_mes,
        resumen
):
    #== == == == == == == == == == == == =
    # LOGO
    # =========================

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte mensual"

    # =========================
    # LOGO
    # =========================

    logo_path = os.path.join(
        os.getcwd(),
        "app",
        "static",
        "img",
        "logocorpo.png"
    )

    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.height = 80
        logo.width = 220
        ws.add_image(logo, "A1")
    ws.title = "Reporte mensual"

    # =========================
    # ESTILOS
    # =========================

    titulo_font = Font(size=18, bold=True)
    header_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font_white = Font(bold=True, color="FFFFFF")

    # =========================
    # TITULO
    # =========================

    ws.merge_cells("A3:D3")
    ws["A3"] = "Reporte mensual de asistencia"
    ws["A3"].font = titulo_font
    ws["A3"].alignment = center

    ws.merge_cells("A2:D2")
    ws["A2"] = empresa_nombre
    ws["A2"].alignment = center

    ws.merge_cells("A3:D3")
    ws["A3"] = nombre_mes
    ws["A3"].alignment = center

    # =========================
    # ENCABEZADOS
    # =========================

    headers = [
        "Empleado",
        "Días trabajados",
        "Horas trabajadas",
        "Estado"
    ]

    row_start = 7

    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=row_start, column=col)

        cell.value = header
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center

    # =========================
    # DATOS
    # =========================

    row = row_start + 1

    for r in resumen:

        ws.cell(row=row, column=1).value = \
            f"{r['empleado'].apellido}, {r['empleado'].nombre}"

        ws.cell(row=row, column=2).value = r["dias"]

        ws.cell(row=row, column=3).value = r["horas"]

        estado_cell = ws.cell(row=row, column=4)
        estado_cell.value = r["estado"]

        if r["estado"] == "INCOMPLETO":
            estado_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        else:
            estado_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

        row += 1

    # =========================
    # TOTALES
    # =========================
    ws.cell(row=row + 1, column=1).value = "Total empleados"
    ws.cell(row=row + 1, column=2).value = len(resumen)
    total_horas = 0
    total_minutos = 0
    for r in resumen:
        h, m = r["horas"].split(":")
        total_horas += int(h)
        total_minutos += int(m)
    extra_horas = total_minutos // 60
    total_horas += extra_horas
    total_minutos = total_minutos % 60
    ws.cell(row=row + 2, column=1).value = "Total horas"
    ws.cell(row=row + 2, column=2).value = f"{total_horas:02d}:{total_minutos:02d}"

    # =========================
    # AJUSTAR COLUMNAS
    # =========================

    widths = [35, 20, 20, 15]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # =========================
    # GUARDAR EN MEMORIA
    # =========================

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return file

#------------------------------------------------
#REPORTE DETALLE POR EMPLEADO
#------------------------------------------------
def exportar_detalle_empleado_excel(
        empresa_nombre,
        empleado_nombre,
        nombre_mes,
        detalle
):

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle mensual"

    titulo_font = Font(size=18, bold=True)
    header_font = Font(bold=True, color="FFFFFF")

    center = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    # TITULO
    ws.merge_cells("A1:G1")
    ws["A1"] = "Detalle mensual de asistencia"
    ws["A1"].font = titulo_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = empresa_nombre
    ws["A2"].alignment = center

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Empleado: {empleado_nombre}"
    ws["A3"].alignment = center

    ws.merge_cells("A4:G4")
    ws["A4"] = nombre_mes
    ws["A4"].alignment = center

    headers = [
        "Fecha",
        "Bloque",
        "Ingreso",
        "Salida",
        "Horas",
        "Estado",
        "Actividad"
    ]

    row_start = 6

    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=row_start, column=col)

        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = row_start + 1

    for d in detalle:

        ws.cell(row=row, column=1).value = d["fecha"].strftime("%d/%m")
        ws.cell(row=row, column=2).value = d["bloque"]

        ws.cell(row=row, column=3).value = (
            d["ingreso"].strftime("%H:%M")
            if d["ingreso"] else "-"
        )

        ws.cell(row=row, column=4).value = (
            d["salida"].strftime("%H:%M")
            if d["salida"] else "-"
        )

        ws.cell(row=row, column=5).value = d["horas"]
        ws.cell(row=row, column=6).value = d["estado"]
        ws.cell(row=row, column=7).value = d["actividad"]

        row += 1

    widths = [12,10,12,12,12,12,20]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return file